###############################################################################################
#  
#  Created by: Jonathan Tan (jona1115@iastate.edu)
#  
###########################################################################################
#  
#  ExcelCreaterFromDict.py:
#  For creating Excel file using python dictionaries. This file is created with, or with the
#  aid of, ChatGPT.
#  
#  MODIFICATION HISTORY:
# 
#  Ver   Who       Date	      Changes
#  ----- --------- ---------- ----------------------------------------------
#  1.00	 Jonathan  5/14/2024  Created file.
#  
###############################################################################################

import openpyxl
from openpyxl.styles import Font

def create_excel_from_dicts(dicts, filename):
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    col = 1  # Start from the first column

    for d in dicts:
        # Write the keys
        row = 1
        for key in d.keys():
            ws.cell(row=row, column=col, value=key).font = Font(bold=True)
            row += 1
        
        # Write the values
        row = 1
        for value in d.values():
            ws.cell(row=row, column=col + 1, value=value)
            row += 1

        col += 3  # Move to the next pair of columns (1 for space + 2 for keys and values)

    # Save the workbook
    wb.save(filename)

def create_excel_from_dicts_with_titles(dicts, titles, filename):
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    col = 1  # Start from the first column

    for i, d in enumerate(dicts):
        # Write the title
        title_cell = ws.cell(row=1, column=col, value=titles[i])
        title_cell.font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        
        # Write the keys
        row = 2  # Start from the second row to leave space for the title
        for key in d.keys():
            key_cell = ws.cell(row=row, column=col, value=key)
            key_cell.font = Font(bold=True)
            row += 1
        
        # Write the values
        row = 2  # Start from the second row to leave space for the title
        for value in d.values():
            ws.cell(row=row, column=col + 1, value=value)
            row += 1

        col += 3  # Move to the next pair of columns (1 for space + 2 for keys and values)

    # Save the workbook
    wb.save(filename)

def append_row_with_text(filename, row_data, bold_columns=[]):
    """
    Append a row to the Excel sheet with the specified text.
    
    :param filename: The name of the Excel file.
    :param row_data: A list of values to be added in the new row.
    :param bold_columns: A list of column indices (1-based) to be bold.
    """
    # Load the workbook and select the active worksheet
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    # Determine the next row index
    next_row = ws.max_row + 1
    
    # Append the row data
    for col_index, value in enumerate(row_data, start=1):
        cell = ws.cell(row=next_row, column=col_index, value=value)
        if col_index in bold_columns:
            cell.font = Font(bold=True)
    
    # Save the workbook
    wb.save(filename)

def prepend_row_with_text(filename, row_data, bold_columns=[]):
    """
    Prepend a row to the Excel sheet with the specified text.
    
    :param filename: The name of the Excel file.
    :param row_data: A list of values to be added in the new row.
    :param bold_columns: A list of column indices (1-based) to be bold.
    """
    # Load the workbook and select the active worksheet
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    # Insert a blank row at the top
    ws.insert_rows(1)
    
    # Add the new row data
    for col_index, value in enumerate(row_data, start=1):
        cell = ws.cell(row=1, column=col_index, value=value)
        if col_index in bold_columns:
            cell.font = Font(bold=True)
    
    # Save the workbook
    wb.save(filename)


if __name__ == "__main__":
    # # Example usage
    # dict1 = {'a': 1, 'b': 2, 'c': 3}
    # dict2 = {'x': 10, 'y': 20, 'z': 30}
    # dict3 = {'i': 22, 'j': 44, 'k': 88}
    # dicts = [dict1, dict2, dict3]

    # create_excel_from_dicts(dicts, './collecteddata/test.xlsx')

    # Example usage
    dict1 = {'a': 1, 'b': 2, 'c': 3}
    dict2 = {'x': 10, 'y': 20, 'z': 30}
    dicts = [dict1, dict2]
    titles = ["Dictionary 1", "Dictionary 2"]

    create_excel_from_dicts_with_titles(dicts, titles, './collecteddata/output_with_titles.xlsx')
